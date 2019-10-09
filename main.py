# -*- coding: utf-8 -*-
import argparse
import asyncio
import datetime
import os
import time
from collections import namedtuple
from io import StringIO

import aiohttp
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from config import codes_path, send_tos, template_path, template_back_path, send_to_myself
from mail import send_mail

url = 'http://stock.gtimg.cn/data/index.php'

params = {
    "appn": "detail",
    "action": "download",
    "c": "sz000004",
    "d": "20190806"
}

Code = namedtuple('Code', 'name code amount volume type')

map_type_to_num = {"卖盘": 3,
                   "中性盘": 2,
                   "买盘": 1}


def now():
    return time.time()


def download(date, all_codes):
    params["d"] = date
    semaphore = asyncio.Semaphore(200)

    async def download_one(code):
        params["c"] = code
        async with aiohttp.ClientSession() as session:
            async with session.get(url, params=params) as resp:
                if resp and resp.status == 200:
                    data = await resp.text(encoding="gbk")
                    return data
                else:
                    print(f"{code} 下载失败")
                    return

    async def process_one(single):
        async with semaphore:
            data = await download_one(single.code)
            if data and data != "暂无数据":
                df = pd.read_csv(StringIO(data), sep="\t",
                                 names=['time', 'price', 'change', 'volume', 'amount', 'type'],
                                 skiprows=[0])
                if df.shape[0] == 0:
                    print(f"{single.code} 当日无交易")
                    return
                volume = df.iloc[-1]["volume"]
                _type = df.iloc[-1]["type"]
                amount = df.amount.sum()
                single_code = Code(name=single.name, volume=volume, amount=amount, code=single.code, type=_type)
                df.to_excel(f"data/{date}/{single.code}.xls")
                return single_code

    tasks = [asyncio.ensure_future(process_one(single)) for single in all_codes]
    event_loop = asyncio.get_event_loop()
    event_loop.run_until_complete(asyncio.wait(tasks))
    codes = [task.result() for task in tasks]
    codes = list(filter(lambda x: x is not None, codes))

    return codes


def condition(code_data):
    """
    全天无901以上买单或者卖单条件限定
    :param code_data: dataframe
    :return: Boolean
    """
    condition1 = code_data[
        (code_data["volume"] >= 901) & (code_data["type"] == "买盘")
        & (code_data["time"] > datetime.time(9, 32)) & (code_data["time"] < datetime.time(14, 30))]

    condition2 = code_data[
        (code_data["volume"] >= 901) & (code_data["type"] == "卖盘")
        & (code_data["time"] > datetime.time(9, 32)) & (code_data["time"] < datetime.time(14, 57))]

    if condition1.empty and condition2.empty:
        return True
    else:
        return False


def run(date, over=1000, below=10000):
    # 创建存储文件夹
    if not os.path.exists("data"):
        os.makedirs("data")
    if not os.path.exists("result"):
        os.makedirs("result")

    # 存储情况1~7的code
    column1 = []
    column2 = []
    column3 = []
    column4 = []
    column5 = []
    column6 = []
    column7 = []

    # 读取待查询股票
    all_data = pd.read_excel(codes_path)
    all_data.columns = ["code", "name"]
    all_data.code = all_data.code.str.replace("SZ", "sz").str.replace("SH", "sh")

    # 下载所有数据
    download_start = now()
    codes_data = download(date=date, all_codes=all_data.itertuples())
    download_end = now()
    print(f"下载用时 {download_end - download_start} s")
    codes_data.sort(key=lambda x: (x.volume, map_type_to_num[x.type]), reverse=True)

    # 情况5,6 现手300以上 情况5 成交金额大于1600万
    for single_code in codes_data:
        if not (single_code.volume > 300 and single_code.type == "卖盘"):
            continue

        # 读取文件，处理时间
        code_data = pd.read_excel(f"data/{date}/{single_code.code}.xls")
        code_data['time'] = pd.to_datetime(code_data['time'], format="%H:%M:%S")
        code_data['time'] = code_data['time'].dt.time

        # 全天无901以上买单或者卖单条件限定
        if not condition(code_data):
            continue

        # 情况5：9点40分前无101以上买单或者卖单，10点前无101以上卖单，且全天无901以上买单或者卖单，现手300以上，成交金额大于1600万
        condition1 = code_data[
            (code_data["volume"] >= 101) & (code_data["type"].isin(["买盘", "卖盘"]))
            & (code_data["time"] > datetime.time(9, 30)) & (code_data["time"] < datetime.time(9, 40))]

        condition2 = code_data[
            (code_data["volume"] >= 101) & (code_data["type"] == "卖盘")
            & (code_data["time"] > datetime.time(9, 30)) & (code_data["time"] < datetime.time(10, 0))]

        if condition1.empty and condition2.empty and single_code.amount > 1600_0000:
            column5.append(single_code)

        # 情况6：9点40分前有白单（大于等于10），其后有白单（大于等于10），其后有白单（大于等于10），且全天无901以上买单或者卖单，现手300以上
        condition1 = code_data[
            (code_data["volume"] >= 10) & (code_data["type"] == "中性盘")
            & (code_data["time"] > datetime.time(9, 30)) & (code_data["time"] < datetime.time(9, 40))]
        condition1_index = list(condition1.index)
        for_query = code_data[
            (code_data["volume"] >= 10) & (code_data["type"] == "中性盘")
            & (code_data["time"] > datetime.time(9, 30))]
        for_query_index = list(for_query.index)
        added = False
        for i in condition1_index:
            if len(set(for_query_index) & {i, i + 1, i + 2}) == 3:
                added = True
                break
        if added:
            column6.append(single_code)

    codes_data_filter = filter(lambda x: x.type == "卖盘" and below >= x.volume >= over, codes_data)

    # 情况1-6筛选
    for single_code in codes_data_filter:
        # 读取文件，处理时间
        code_data = pd.read_excel(f"data/{date}/{single_code.code}.xls")
        code_data['time'] = pd.to_datetime(code_data['time'], format="%H:%M:%S")
        code_data['time'] = code_data['time'].dt.time

        # 情况1：9点40分前有上万白单，全天无大于10001的卖单或者卖单
        condition1 = code_data[
            (code_data["volume"] >= 10000) & (code_data["type"] == "中性盘")
            & (code_data["time"] > datetime.time(9, 30)) & (code_data["time"] < datetime.time(9, 40))]

        condition2 = code_data[
            (code_data["volume"] > 10001) & (code_data["type"].isin(["买盘", "卖盘"]))
            & (code_data["time"] > datetime.time(9, 32)) & (code_data["time"] < datetime.time(14, 57))]

        if (not condition1.empty) and condition2.empty:
            column1.append(single_code)
            continue

        # 情况2：9点40分前有连续上千白单，全天无大于9001的买单或者卖单
        condition1 = code_data[
            (code_data["volume"] >= 1000) & (code_data["type"] == "中性盘")
            & (code_data["time"] > datetime.time(9, 30)) & (code_data["time"] < datetime.time(9, 40))]
        condition1_index = list(condition1.index)

        condition2 = code_data[
            (code_data["volume"] > 9001) & (code_data["type"].isin(["买盘", "卖盘"]))
            & (code_data["time"] > datetime.time(9, 32)) & (code_data["time"] < datetime.time(14, 57))]

        if len(condition1_index) >= 2 and condition2.empty:
            delta = [condition1_index[i + 1] - condition1_index[i] for i in range(len(condition1_index) - 1)]
            if min(delta) <= 6:
                column2.append(single_code)
                continue

        # 全天无901以上买单或者卖单条件限定
        if not condition(code_data):
            continue

        # 情况3：9点40分前有上千白单，全天无901以上买单或者卖单
        condition1 = code_data[
            (code_data["volume"] >= 1000) & (code_data["type"] == "中性盘")
            & (code_data["time"] > datetime.time(9, 30)) & (code_data["time"] < datetime.time(9, 40))]

        if not condition1.empty:
            column3.append(single_code)
            continue

        # 情况4：9点40分前有连续白单（均大于100），全天无901以上买单或者卖单
        condition1 = code_data[
            (code_data["volume"] > 100) & (code_data["type"] == "中性盘")
            & (code_data["time"] > datetime.time(9, 30)) & (code_data["time"] < datetime.time(9, 40))]

        condition1_index = list(condition1.index)

        if len(condition1_index) >= 2:
            delta = [condition1_index[i + 1] - condition1_index[i] for i in range(len(condition1_index) - 1)]
            if min(delta) <= 6:
                column4.append(single_code)
                continue

        # 情况7：9点40分前有白单（大于100，小于1000），且全天无901以上买单或者卖单
        condition1 = code_data[
            (code_data["volume"] > 100) & (code_data["volume"] < 1000) & (code_data["type"] == "中性盘")
            & (code_data["time"] > datetime.time(9, 30)) & (code_data["time"] < datetime.time(9, 40))]

        if not condition1.empty:
            column7.append(single_code)
            continue

    # 数据汇总
    final_data = [column1, column2, column3, column4, column5, column6, column7]
    print(f"情况 1~7 符合条件数目 {[len(item) for item in final_data]}")
    final_name_data = [[i.name for i in j] for j in final_data]
    final_annotation_data = [[f"{i.type}-{i.volume}手\n总交易额{int(i.amount / 10000)}万" for i in j] for j in final_data]
    final_verify_data = [[f"{i.name}-{i.code}\n{i.type}-{i.volume}-{i.amount}" for i in j] for j in final_data]

    # 写出
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    for i, column in enumerate(final_name_data):
        for j, string in enumerate(column):
            ws.cell(row=j + 2, column=2 * i + 1).value = string

    for i, column in enumerate(final_annotation_data):
        for j, string in enumerate(column):
            ws.cell(row=j + 2, column=2 * i + 2).alignment = Alignment(wrap_text=True)
            ws.cell(row=j + 2, column=2 * i + 2).value = string

    wb.save(f"result/{date}结果.xlsx")

    wb = load_workbook(template_back_path)
    ws = wb[wb.sheetnames[0]]
    for i, column in enumerate(final_verify_data):
        for j, string in enumerate(column):
            ws.cell(row=j + 2, column=i + 1).alignment = Alignment(wrap_text=True)
            ws.cell(row=j + 2, column=i + 1).value = string

    wb.save(f"result/{date}结果供查验.xlsx")

    return


def is_weekday(date):
    params["d"] = date
    resp = requests.get(url, params=params)
    resp.encoding = "gbk"
    return resp.text != '暂无数据'


def main(date, myself):
    date_print = datetime.datetime.strptime(date, "%Y%m%d").strftime("%Y-%m-%d")
    if is_weekday(date):
        if not os.path.isdir(f"data/{date}"):
            os.makedirs(f"data/{date}")
        start = now()
        print(f"{date_print} 有数据")
        run(date)
        end = now()
        print(f"总用时 {end - start} s")
        text = f"{date_print}\n" \
               f"总用时 {end - start} s\n" \
               f"请查收附件."
        people_send_to = send_to_myself if myself else send_tos
        send_mail(send_tos=people_send_to, name="Simon Yang", subject=f"{date_print}结果", text=text,
                  att_urls=[f"result/{date}结果.xlsx", f"result/{date}结果供查验.xlsx"])
    else:
        print(f"{date_print} 休市 无数据")
        # send_mail(send_tos=send_tos, name="Simon Yang", subject=f"{today_print}结果", text="今天休市，无数据")


if __name__ == '__main__':
    today = datetime.datetime.now().strftime("%Y%m%d")
    parser = argparse.ArgumentParser()
    parser.add_argument("-d", "--date", help="specify the date")
    parser.add_argument("-m", "--myself", help="specify which to send to", action="store_true")
    args = parser.parse_args()
    if args.date:
        main(date=args.date, myself=args.myself)
    else:
        main(date=today, myself=args.myself)
